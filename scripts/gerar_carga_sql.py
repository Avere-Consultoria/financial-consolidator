"""Lê Base_Inicial.xlsx e gera SQL de carga idempotente + backfill + reconciliação."""
import os, csv, re
from openpyxl import load_workbook

SRC = r"C:\Users\luiz_\Downloads\Base_Inicial.xlsx"
OUT = os.path.join(os.path.dirname(__file__), "out")
os.makedirs(OUT, exist_ok=True)

def q(v):
    if v is None: return "NULL"
    return "'" + str(v).replace("'", "''") + "'"

def acc(v):
    """Número de conta como texto limpo (sem .0)."""
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return str(int(v))
    if isinstance(v, int): return str(v)
    s = str(v).strip()
    return s or None

wb = load_workbook(SRC, data_only=True, read_only=True)

# ── Consultores ───────────────────────────────────────────────────────────────
wsc = wb["Consultores"]
crows = [r for r in wsc.iter_rows(values_only=True)][1:]
consultores = []
for r in crows:
    if not r or r[0] in (None, ""): continue
    consultores.append({"codigo": int(r[0]), "nome": str(r[1]).strip(), "email": str(r[2]).strip().lower()})

# ── Clientes + Contas ─────────────────────────────────────────────────────────
wsk = wb["Clientes"]
krows = list(wsk.iter_rows(values_only=True))
hdr = list(krows[0])
APIS = {"Cod_XP1": ("XP", 1), "Cod_XP2": ("XP", 2), "Cod_BTG1": ("BTG", 1), "Cod_BTG2": ("BTG", 2)}
idx = {h: i for i, h in enumerate(hdr)}
manual_cols = hdr[9:]  # ITAU..GALAPAGOS

clientes, contas = [], []
bancos_manuais = set()
for r in krows[1:]:
    if not r or r[0] in (None, ""): continue
    cod = str(int(r[0])) if isinstance(r[0], (int, float)) else str(r[0]).strip()
    nome = str(r[1]).strip()
    direto = int(r[2]) if r[2] not in (None, "") else None
    clientes.append({"cod": cod, "nome": nome, "direto": direto})
    # contas API (XP/BTG)
    for col, (base, ordem) in APIS.items():
        v = acc(r[idx[col]])
        if v: contas.append({"cod": cod, "tipo": "API", "inst": base, "codigo": v, "ordem": ordem})
    # contas manuais
    for col in manual_cols:
        v = acc(r[idx[col]])
        if v:
            bancos_manuais.add(col.strip().upper())
            contas.append({"cod": cod, "tipo": "MANUAL", "inst": col.strip().upper(), "codigo": v, "ordem": 1})

# ── SQL de carga ──────────────────────────────────────────────────────────────
sql = []
sql.append("-- CARGA INICIAL AVERE — idempotente. Rode APÓS a migration 20260618 (codigo_interno).")
sql.append("BEGIN;")
sql.append("")
sql.append("-- Garante chave única para os UPSERTs (no-op se já existir).")
sql.append("CREATE UNIQUE INDEX IF NOT EXISTS uq_clientes_codigo_avere ON clientes(codigo_avere);")
sql.append("")
sql.append("-- ── Consultores (upsert por codigo_interno; preserva perfil_id existente) ──")
for c in consultores:
    sql.append(
        f"INSERT INTO consultores (codigo_interno, nome, email_professional, ativo) "
        f"VALUES ({c['codigo']}, {q(c['nome'])}, {q(c['email'])}, true) "
        f"ON CONFLICT (codigo_interno) WHERE codigo_interno IS NOT NULL "
        f"DO UPDATE SET nome=EXCLUDED.nome, email_professional=EXCLUDED.email_professional;"
    )
sql.append("")
sql.append("-- ── Instituições manuais (cria se não existir) ──")
for b in sorted(bancos_manuais):
    sql.append(
        f"INSERT INTO instituicoes (nome, tipo, cor_primaria) "
        f"SELECT {q(b)}, 'MANUAL', '#64748B' "
        f"WHERE NOT EXISTS (SELECT 1 FROM instituicoes WHERE upper(nome)=upper({q(b)}));"
    )
sql.append("")
sql.append("-- ── Clientes (upsert por codigo_avere; consultor_id via codigo_interno → perfil) ──")
for k in clientes:
    cid = f"(SELECT id FROM consultores WHERE codigo_interno = {k['direto']})" if k["direto"] else "NULL"
    sql.append(
        f"INSERT INTO clientes (codigo_avere, nome, consultor_id) "
        f"VALUES ({q(k['cod'])}, {q(k['nome'])}, {cid}) "
        f"ON CONFLICT (codigo_avere) DO UPDATE SET nome=EXCLUDED.nome, consultor_id=EXCLUDED.consultor_id;"
    )
sql.append("")
sql.append("-- ── Contas (upsert por (instituicao_id, codigo)) ──")
for c in contas:
    if c["tipo"] == "API":
        inst = f"(SELECT id FROM instituicoes WHERE codigo = {q(c['inst'])} LIMIT 1)"
    else:
        inst = f"(SELECT id FROM instituicoes WHERE upper(nome)=upper({q(c['inst'])}) LIMIT 1)"
    sql.append(
        f"INSERT INTO cliente_contas (cliente_id, instituicao_id, codigo, ordem, ativo) "
        f"SELECT (SELECT id FROM clientes WHERE codigo_avere={q(c['cod'])}), {inst}, {q(c['codigo'])}, {c['ordem']}, true "
        f"ON CONFLICT (instituicao_id, codigo) WHERE codigo IS NOT NULL AND codigo <> '' "
        f"DO UPDATE SET cliente_id=EXCLUDED.cliente_id, ordem=EXCLUDED.ordem, updated_at=now();"
    )
sql.append("")
sql.append("COMMIT;")
sql.append("")
sql.append("-- ── Conferência ──")
sql.append("SELECT 'consultores' AS t, count(*) FROM consultores WHERE codigo_interno IS NOT NULL;")
sql.append("SELECT 'clientes' AS t, count(*) FROM clientes;")
sql.append("SELECT 'contas' AS t, count(*) FROM cliente_contas;")
sql.append("SELECT i.nome, i.codigo, count(*) FROM cliente_contas cc JOIN instituicoes i ON i.id=cc.instituicao_id GROUP BY i.nome, i.codigo ORDER BY count(*) DESC;")

with open(os.path.join(OUT, "carga_inicial.sql"), "w", encoding="utf-8") as f:
    f.write("\n".join(sql))

# ── Backfill de vínculo (re-rodável após provisionar logins) ──────────────────
bf = []
bf.append("-- BACKFILL: liga clientes.consultor_id ao perfil do consultor já provisionado.")
bf.append("-- Re-rode quando criar/provisionar logins de consultores. Idempotente.")
pares = ",\n  ".join(f"({q(k['cod'])}, {k['direto']})" for k in clientes if k["direto"])
bf.append("UPDATE clientes c SET consultor_id = co.id")
bf.append("FROM (VALUES")
bf.append("  " + pares)
bf.append(") AS m(codigo_avere, codigo_interno)")
bf.append("JOIN consultores co ON co.codigo_interno = m.codigo_interno")
bf.append("WHERE c.codigo_avere = m.codigo_avere;")
with open(os.path.join(OUT, "backfill_consultor.sql"), "w", encoding="utf-8") as f:
    f.write("\n".join(bf))

# ── Reconciliação CSV ─────────────────────────────────────────────────────────
with open(os.path.join(OUT, "reconciliacao_carga.csv"), "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f)
    w.writerow(["codigo_avere", "nome", "consultor_codigo_interno", "instituicao", "numero_conta", "ordem"])
    cmap = {k["cod"]: k for k in clientes}
    for c in contas:
        k = cmap.get(c["cod"], {})
        w.writerow([c["cod"], k.get("nome", ""), k.get("direto", ""), c["inst"], c["codigo"], c["ordem"]])

print(f"Consultores: {len(consultores)}")
print(f"Clientes:    {len(clientes)}")
print(f"Contas:      {len(contas)}")
print(f"Bancos manuais: {sorted(bancos_manuais)}")
print(f"Arquivos em: {OUT}")
