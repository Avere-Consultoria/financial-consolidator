from openpyxl import load_workbook
wb = load_workbook(r"C:\Users\luiz_\Downloads\Base_Inicial.xlsx", data_only=True, read_only=True)
ws = wb["Clientes"]
rows = list(ws.iter_rows(values_only=True))
hdr = list(rows[0]); data = rows[1:]
data = [r for r in data if r and r[0] not in (None, "")]
print("Clientes (linhas de dado):", len(data))
# contagem de não-vazios por coluna
print("\nPreenchimento por coluna:")
for i, col in enumerate(hdr):
    n = sum(1 for r in data if i < len(r) and r[i] not in (None, "", " "))
    print(f"  {col:>14}: {n}")
# PF x PJ pelo CPF/CNPJ (col index 3)
import re
pf = pj = outros = 0
for r in data:
    doc = re.sub(r"\D", "", str(r[3] or ""))
    if len(doc) == 11: pf += 1
    elif len(doc) == 14: pj += 1
    else: outros += 1
print(f"\nPF(11): {pf}  PJ(14): {pj}  Indefinido: {outros}")
# consultores distintos referenciados em 'Direto'
diretos = sorted({r[2] for r in data if r[2] not in (None,"")})
print(f"\nConsultores referenciados em 'Direto': {diretos}")
