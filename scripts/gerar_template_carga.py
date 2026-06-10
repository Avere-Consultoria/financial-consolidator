"""Gera o template de planilha para a carga inicial do Consolidador Avere."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

AZUL = "0083CB"
HEADER_FILL = PatternFill("solid", fgColor=AZUL)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NOTE_FONT = Font(italic=True, color="9CA3AF", size=10)
EX_FONT = Font(color="6B7280", size=10)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def add_sheet(nome, colunas, exemplos, notas):
    ws = wb.create_sheet(nome)
    # nota no topo
    ws.cell(row=1, column=1, value=notas).font = NOTE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(colunas)))
    # cabeçalho na linha 3
    for c, (col, larg) in enumerate(colunas, start=1):
        cell = ws.cell(row=3, column=c, value=col)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = larg
    # exemplos
    for r, linha in enumerate(exemplos, start=4):
        for c, val in enumerate(linha, start=1):
            cell = ws.cell(row=r, column=c, value=val); cell.font = EX_FONT; cell.border = BORDER
    ws.freeze_panes = "A4"
    return ws

# ── Instruções ────────────────────────────────────────────────────────────────
ws = wb.active; ws.title = "Instruções"
linhas = [
    ("CARGA INICIAL — CONSOLIDADOR AVERE", Font(bold=True, size=14, color=AZUL)),
    ("", None),
    ("Preencha uma aba por bloco. As linhas em cinza são EXEMPLOS — apague antes de enviar.", NOTE_FONT),
    ("Mantenha os nomes das colunas (linha azul) exatamente como estão.", NOTE_FONT),
    ("", None),
    ("ORDEM DE PRIORIDADE (sem os 3 primeiros nada sincroniza):", Font(bold=True, size=11)),
    ("  1. Consultores   →   2. Clientes   →   3. Contas", None),
    ("  4. Emissores e 5. Instituições manuais podem vir depois.", None),
    ("", None),
    ("CHAVES DE LIGAÇÃO:", Font(bold=True, size=11)),
    ("  • Cliente liga ao Consultor pelo 'consultor_codigo_avere'.", None),
    ("  • Conta liga ao Cliente pelo 'cliente_codigo_avere'.", None),
    ("", None),
    ("IMPORTANTE:", Font(bold=True, size=11)),
    ("  • codigo_avere é a CHAVE — não pode repetir nem mudar depois.", None),
    ("  • CPF/CNPJ: pode mandar com ou sem pontuação.", None),
    ("  • Uma conta por linha. Cliente com 3 contas XP = 3 linhas na aba Contas.", None),
    ("  • Dado sensível: envie por canal seguro.", None),
    ("  • NÃO precisamos de telefone, fee, taxa, nascimento nesta fase (ficam pro CRM).", NOTE_FONT),
]
for i, (txt, font) in enumerate(linhas, start=1):
    cell = ws.cell(row=i, column=1, value=txt)
    if font: cell.font = font
ws.column_dimensions["A"].width = 95

# ── Consultores ───────────────────────────────────────────────────────────────
add_sheet(
    "Consultores",
    [("codigo_avere_consultor", 22), ("nome", 32), ("email_profissional", 34), ("papel", 16)],
    [
        ["1023", "Bruno Silva", "bruno@averepartners.com.br", "CONSULTOR"],
        ["1001", "Maria Diretora", "maria@averepartners.com.br", "MASTER"],
    ],
    "CONSULTORES — equipe que terá acesso ao sistema. papel = CONSULTOR ou MASTER.",
)

# ── Clientes ──────────────────────────────────────────────────────────────────
ws_cli = add_sheet(
    "Clientes",
    [("codigo_avere", 16), ("nome", 34), ("tipo", 10), ("cpf_cnpj", 22), ("consultor_codigo_avere", 24), ("ativo", 10)],
    [
        ["230134", "Carlos Eduardo Souza", "PF", "123.456.789-00", "1023", "SIM"],
        ["230210", "Empresa Acme Ltda", "PJ", "12.345.678/0001-90", "1023", "SIM"],
    ],
    "CLIENTES — tipo = PF ou PJ. consultor_codigo_avere = o código da aba Consultores. ativo = SIM/NAO.",
)
dv_tipo = DataValidation(type="list", formula1='"PF,PJ"', allow_blank=False)
dv_ativo = DataValidation(type="list", formula1='"SIM,NAO"', allow_blank=False)
ws_cli.add_data_validation(dv_tipo); ws_cli.add_data_validation(dv_ativo)
dv_tipo.add("C4:C1000"); dv_ativo.add("F4:F1000")

# ── Contas ────────────────────────────────────────────────────────────────────
ws_con = add_sheet(
    "Contas",
    [("cliente_codigo_avere", 22), ("instituicao", 18), ("numero_conta", 20), ("documento_titular", 22), ("apelido", 24), ("ordem", 8)],
    [
        ["230134", "BTG", "4567890", "123.456.789-00", "", "1"],
        ["230134", "XP", "112233", "123.456.789-00", "XP Aposentadoria", "1"],
        ["230134", "XP", "445566", "123.456.789-00", "XP Trading", "2"],
        ["230134", "AGORA", "778899", "123.456.789-00", "", "1"],
        ["230210", "ITAU", "0001-9", "12.345.678/0001-90", "", "1"],
    ],
    "CONTAS — uma linha por conta. instituicao: BTG / XP / AVENUE / AGORA ou o nome do banco (manual). "
    "numero_conta = identificador na corretora. documento_titular: CPF/CNPJ (Ágora usa CPF). ordem: 1,2,3 se houver +de 1 conta na mesma instituição.",
)
dv_inst = DataValidation(type="list", formula1='"BTG,XP,AVENUE,AGORA,MANUAL"', allow_blank=True)
ws_con.add_data_validation(dv_inst); dv_inst.add("B4:B1000")

# ── Emissores (opcional) ──────────────────────────────────────────────────────
add_sheet(
    "Emissores (opcional)",
    [("nome_fantasia", 34), ("cnpj", 22), ("setor", 24)],
    [
        ["Engie Brasil", "02.474.103/0001-19", "Energia"],
        ["Rumo Logística", "02.387.241/0001-60", "Logística"],
    ],
    "EMISSORES (opcional) — só crédito privado (debêntures, CRA, CRI...). Acelera a classificação. Bancos NÃO precisam (vêm do FGC/BCB).",
)

# ── Instituições manuais (opcional) ───────────────────────────────────────────
add_sheet(
    "Instituicoes_Manuais (opc)",
    [("nome_banco", 28), ("observacao", 50)],
    [
        ["Itaú", "Sem API — posições virão por extrato (PDF/Excel)"],
        ["Banco do Brasil", "Sem API — posições virão por extrato (PDF/Excel)"],
    ],
    "INSTITUIÇÕES MANUAIS (opcional) — bancos/corretoras sem API. Os extratos/posições desses virão à parte (PDF/Excel).",
)

import os
out = os.path.join(os.path.dirname(__file__), "..", "carga_inicial_avere_TEMPLATE.xlsx")
out = os.path.abspath(out)
wb.save(out)
print("OK:", out)
