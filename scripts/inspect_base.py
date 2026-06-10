from openpyxl import load_workbook
wb = load_workbook(r"C:\Users\luiz_\Downloads\Base_Inicial.xlsx", data_only=True, read_only=True)
for ws in wb.worksheets:
    print("="*80)
    print(f"ABA: {ws.title}  (linhas~{ws.max_row}, colunas~{ws.max_column})")
    rows = list(ws.iter_rows(values_only=True))
    # acha a linha de cabeçalho (primeira não-vazia)
    for i, r in enumerate(rows[:8]):
        print(f"  L{i}: {list(r)}")
    print(f"  ...total de linhas lidas: {len(rows)}")
